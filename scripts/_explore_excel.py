# -*- coding: utf-8 -*-
"""一次性探索腳本：拆解 teacher.xlsx 的 summary sheet 結構。
跑完後刪除即可。"""
import openpyxl
from openpyxl.utils import get_column_letter
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

wb_v = openpyxl.load_workbook('data/teacher.xlsx', data_only=True)
wb_f = openpyxl.load_workbook('data/teacher.xlsx', data_only=False)

summary_name = wb_v.sheetnames[0]
ws = wb_v[summary_name]
wsf = wb_f[summary_name]

print(f"summary sheet: {summary_name!r}")
print(f"dims: {ws.max_row} rows x {ws.max_column} cols\n")

# Row 4 position coefficient headers
print("=== Row 4 headers (職務節數係數列) ===")
for c in range(22, 38):
    label = ws.cell(row=3, column=c).value
    val = ws.cell(row=4, column=c).value
    print(f"  {get_column_letter(c)}3={label!r}  {get_column_letter(c)}4={val!r}")

# Domain rows scan
print("\n=== Domain rows (all summary B/C/D/G/I/O/L/M/Q/AH/AI/AK) ===")
for r in range(5, 20):
    b = ws.cell(row=r, column=2).value
    c_val = ws.cell(row=r, column=3).value
    d_val = ws.cell(row=r, column=4).value
    e_val = ws.cell(row=r, column=5).value
    g_val = ws.cell(row=r, column=7).value
    i_val = ws.cell(row=r, column=9).value
    k_val = ws.cell(row=r, column=11).value
    l_val = ws.cell(row=r, column=12).value
    m_val = ws.cell(row=r, column=13).value
    o_val = ws.cell(row=r, column=15).value
    p_val = ws.cell(row=r, column=16).value
    q_val = ws.cell(row=r, column=17).value
    ah_val = ws.cell(row=r, column=34).value
    ai_val = ws.cell(row=r, column=35).value
    ak_val = ws.cell(row=r, column=37).value
    print(f"  r{r} {b!r}")
    print(f"    113: C={c_val} D={d_val} E={e_val}  G={g_val} H={d_val} I={i_val}")
    print(f"    114: K={k_val} L={l_val} M={m_val}  O={o_val} P={p_val} Q={q_val}")
    print(f"    cnt: AH(正式)={ah_val} AI(加權)={ai_val} AK(代理)={ak_val}")

# Extract formulas
print("\n=== Summary formulas per domain row ===")
for r in range(5, 17):
    b = ws.cell(row=r, column=2).value
    print(f"\n  r{r} {b!r}")
    for col_letter, col_idx in [('C',3),('D',4),('G',7),('K',11),('L',12),('O',15)]:
        f = wsf.cell(row=r, column=col_idx).value
        print(f"    {col_letter}{r} formula = {f!r}")

# Extract W..AG per-position counts for each domain
print("\n=== Per-domain position counts (W..AG) ===")
position_labels_row3 = {}
position_vals_row4 = {}
for c in range(23, 34):  # W..AG
    position_labels_row3[get_column_letter(c)] = ws.cell(row=3, column=c).value
    position_vals_row4[get_column_letter(c)] = ws.cell(row=4, column=c).value
print(f"  labels (row 3): {position_labels_row3}")
print(f"  rates  (row 4): {position_vals_row4}")

for r in range(5, 17):
    b = ws.cell(row=r, column=2).value
    counts = {}
    for c in range(23, 34):
        v = ws.cell(row=r, column=c).value
        if v:
            counts[get_column_letter(c)] = v
    print(f"  r{r} {b!r}: {counts}")

# Evening school sheet
evening_sheet = wb_v.sheetnames[4]
print(f"\n=== Evening sheet [{4}]: {evening_sheet!r} ===")
ws_e = wb_v[evening_sheet]
wsf_e = wb_f[evening_sheet]
print(f"  dims: {ws_e.max_row} x {ws_e.max_column}")
for r in range(1, min(ws_e.max_row+1, 30)):
    for c in range(1, min(ws_e.max_column+1, 15)):
        v = ws_e.cell(row=r, column=c).value
        if v is not None:
            print(f"    {get_column_letter(c)}{r} = {v!r}")
    print("  ---")
