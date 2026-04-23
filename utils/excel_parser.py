# -*- coding: utf-8 -*-
"""把學校原始 teacher.xlsx 萃取成結構化 school_data.json。

Excel 是寬表、多年度混排、公式散落，不適合直接餵給網頁。這個模組做的事：
1. 讀 summary sheet（第 0 張）的各領域列 (row 5..16)
2. 抽出 W..AG 的職務人數、AK 的代理人數、AL/AO 的備註
3. 抽出 C/G 的公式字串（保留作為每個領域的計算規則）
4. 抽出 data_only 的計算值作為 _excel_raw（驗收基準）
"""

import json
import os
import re
from typing import Any

import openpyxl

SUMMARY_SHEET_INDEX = 0
DAY_113_SHEET_INDEX = 1
DAY_114_SHEET_INDEX = 2
EVENING_113_SHEET_INDEX = 4
EVENING_BASELINE_SHEET_INDEX = 5

POSITION_COLUMNS = [
    {"key": "director",      "col": "W",  "label": "主任",              "rate": 1},
    {"key": "chief_5",       "col": "X",  "label": "組長(5節)",         "rate": 5},
    {"key": "chief_7",       "col": "Y",  "label": "組長(7節)",         "rate": 7},
    {"key": "dept_head_6",   "col": "Z",  "label": "科主任(6節)",       "rate": 6},
    {"key": "dept_head_7",   "col": "AA", "label": "科主任(7節)",       "rate": 7},
    {"key": "homeroom_10",   "col": "AB", "label": "導師/資源教室(10節)","rate": 10},
    {"key": "homeroom_12",   "col": "AC", "label": "導師(12節)",        "rate": 12},
    {"key": "homeroom_14",   "col": "AD", "label": "特教(14節)",        "rate": 14},
    {"key": "teacher_12",    "col": "AE", "label": "專任(12節)",        "rate": 12},
    {"key": "teacher_14",    "col": "AF", "label": "專任(14節)",        "rate": 14},
    {"key": "teacher_16",    "col": "AG", "label": "專任(16節)",        "rate": 16},
]

DOMAIN_IDS = {
    5:  ("chinese_social",   "國文/社會領域"),
    6:  ("english",          "英文暨應英科"),
    7:  ("math",             "數學領域"),
    8:  ("science",          "自然領域"),
    9:  ("info_tech",        "資處科"),
    10: ("accounting",       "會計科"),
    11: ("commerce",         "商經科"),
    12: ("multimedia",       "多媒科"),
    13: ("art",              "藝術生活/美術"),
    14: ("pe",               "體育"),
    15: ("health_career",    "健康與護理/生涯規劃"),
    16: ("national_defense", "全民國防教育"),
}


# 日校課程節數表每年的欄位配置（因為 114 多 2 個班，整排向右平移）
# weighted_up / weighted_down：科目列，上/下學期各班加權節數
# sum_up / sum_down：領域總行，對整個領域 range 做 SUM
DAY_SHEET_LAYOUTS = {
    "113": {
        "max_class_col":    31,  # AE
        "weighted_up_col":  32,  # AF = 上學期 subject 總和
        "weighted_down_col":33,  # AG = 下學期 subject 總和
        "sum_up_col":       34,  # AH = 上學期 domain 總和
        "sum_down_col":     35,  # AI = 下學期 domain 總和
    },
    "114": {
        "max_class_col":    33,  # AG
        "weighted_up_col":  34,  # AH
        "weighted_down_col":35,  # AI
        "sum_up_col":       36,  # AJ
        "sum_down_col":     37,  # AK
    },
}


def _extract_formula(cell_value: Any) -> str | None:
    if cell_value is None:
        return None
    if isinstance(cell_value, str) and cell_value.startswith("="):
        return cell_value[1:]
    return None


def _safe_int(v: Any) -> int:
    if v is None:
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


_AI_TAIL = re.compile(r"\)\s*([+\-]\s*\d+)\s*$")

# 科目列加權公式 term：'F5*2' / 'D5' / 'X5*3'
_WEIGHTED_TERM = re.compile(r"^([A-Z]+)\d+(?:\s*\*\s*(\d+))?$")

# 領域 SUM 公式：'SUM(AF5:AF13)' / 'SUM(AF124)' 可帶尾巴 '-2'
_SUM_RANGE = re.compile(
    r"SUM\s*\(\s*([A-Z]+)(\d+)(?:\s*:\s*([A-Z]+)(\d+))?\s*\)\s*([+\-]\s*\d+)?"
)

# 顯式 cell 加法：'AG43+AG42' 風格（例：113 science 下學期 AI41 = AG43+AG42）
_CELL_REF_ONLY = re.compile(r"^([A-Z]+)(\d+)$")


def _parse_weighted_formula(formula: str | None) -> dict | None:
    """把 'F5*2+H5*2+X5*3+D5' 拆成 {'coef': {'F':2,'H':2,'X':3,'D':1}, 'const': 0}。

    各 col 係數代表該班的實際開課數（例：廣告科同時 2 班都修，係數 2；部分科目高三 1 班則 1）。
    編輯 cell 時 subject 加權節數 = Σ cells[col] × coef[col] + const。
    """
    if not formula:
        return None
    body = formula[1:] if formula.startswith("=") else formula
    coef: dict[str, int] = {}
    const = 0
    for tok in body.split("+"):
        tok = tok.strip()
        if not tok:
            continue
        m = _WEIGHTED_TERM.match(tok)
        if m:
            col = m.group(1)
            c = int(m.group(2)) if m.group(2) else 1
            coef[col] = c
            continue
        try:
            const += int(tok)
        except ValueError:
            pass
    return {"coef": coef, "const": const}


def _parse_sum_range(formula: str | None) -> dict | None:
    """把 'SUM(AF5:AF13)-2' 拆成 {'col': 'AF', 'rows': [5..13], 'start': 5, 'end': 13, 'tail': -2}。

    支援三種格式：
      - 'SUM(AF5:AF13)'    範圍型（rows = [start..end]）
      - 'SUM(AF124)'       單格型（rows = [124]）
      - 'AG43+AG42'        顯式 cell 加法（例：113 science 下學期 AI41；rows 依出現順序）
    任一格式都可帶尾巴 '±N'。rows 為實際參與加總的列清單；start/end 為 min/max，供顯示。
    """
    if not formula:
        return None
    body = formula[1:] if formula.startswith("=") else formula

    m = _SUM_RANGE.search(body)
    if m:
        col_from, row_from, col_to, row_to, tail = m.groups()
        if col_to is None:
            col_to = col_from
            row_to = row_from
        if col_from != col_to:
            return None
        start = int(row_from)
        end = int(row_to)
        return {
            "col":   col_from,
            "rows":  list(range(start, end + 1)),
            "start": start,
            "end":   end,
            "tail":  int(tail.replace(" ", "")) if tail else 0,
        }

    # 退而解析顯式 cell 加法
    tokens = [t.strip() for t in body.split("+") if t.strip()]
    if not tokens:
        return None
    col = None
    rows: list[int] = []
    tail = 0
    for tok in tokens:
        m2 = _CELL_REF_ONLY.match(tok)
        if m2:
            c, r = m2.group(1), int(m2.group(2))
            if col is None:
                col = c
            elif c != col:
                return None
            rows.append(r)
            continue
        try:
            tail += int(tok)
        except ValueError:
            return None
    if col is None or not rows:
        return None
    return {
        "col":   col,
        "rows":  rows,
        "start": min(rows),
        "end":   max(rows),
        "tail":  tail,
    }


def _parse_ai_adjustment(ai_formula: str | None) -> int:
    """Excel 原表 AI 公式可能有尾部調整（例：資處科 AI=(... )-7）。抽出這個 ±N 常數。"""
    if not ai_formula:
        return 0
    m = _AI_TAIL.search(ai_formula)
    if not m:
        return 0
    return int(m.group(1).replace(" ", ""))


def _extract_ah_override(ah_cell_value: Any, ah_formula: str | None) -> int | None:
    """若 AH 是 SUM 公式就回 None（要由 positions 加總），若是硬編碼數字就回該值。"""
    if ah_formula and ah_formula.upper().replace(" ", "").startswith("SUM("):
        return None
    if isinstance(ah_cell_value, (int, float)):
        return int(ah_cell_value)
    return None


def _extract_ai_override(ai_cell_value: Any, ai_formula: str | None) -> int | None:
    """AI 若是硬編碼數字（如多媒科 AI=20）就回該值。"""
    if ai_formula:
        return None
    if isinstance(ai_cell_value, (int, float)):
        return int(ai_cell_value)
    return None


_CELL_REF = re.compile(r"([A-Z]+)(\d+)")


def _resolve_cell_value(wb_v: Any, sheet_name: str, cell_ref: str) -> int:
    """取 wb_v（data_only）的指定 cell 整數值；空或非數字回 0。"""
    m = _CELL_REF.match(cell_ref.strip())
    if not m:
        return 0
    col = openpyxl.utils.column_index_from_string(m.group(1))
    row = int(m.group(2))
    try:
        ws = wb_v[sheet_name]
    except KeyError:
        return 0
    v = ws.cell(row=row, column=col).value
    if isinstance(v, (int, float)):
        return int(v)
    return 0


def _decompose_required_formula(
    formula: str | None,
    wb_v: Any,
    day_sheet_name: str,
    evening_sheet_name: str,
) -> dict:
    """把 sheet0 D/L 欄需求公式拆成 day_hours + evening_hours。

    典型：'=日113課程節數預估表!AH5-9+進113課程節數一覽表!K4'
         → day=AH5-9=103, evening=K4=29, total=132
    '=9+進113課程節數一覽表!J73' → day=9, evening=J73
    """
    if not formula:
        return {"day": 0, "evening": 0, "total": 0, "source": None}
    s = formula[1:] if formula.startswith("=") else formula
    tokens = re.split(r"([+\-])", s)
    sign = 1
    day_val = 0
    eve_val = 0
    for tok in tokens:
        tok = tok.strip()
        if tok == "+":
            sign = 1
            continue
        if tok == "-":
            sign = -1
            continue
        if not tok:
            continue
        if "!" in tok:
            sheet_part, cell_part = tok.split("!", 1)
            sheet_part = sheet_part.strip().strip("'")
            v = _resolve_cell_value(wb_v, sheet_part, cell_part)
            if sheet_part == day_sheet_name:
                day_val += sign * v
            elif sheet_part == evening_sheet_name:
                eve_val += sign * v
        else:
            try:
                day_val += sign * int(tok)
            except ValueError:
                pass
    return {
        "day": day_val,
        "evening": eve_val,
        "total": day_val + eve_val,
        "source": formula,
    }


def _extract_day_schedule(
    ws_v: Any,
    ws_f: Any,
    layout: dict,
    summary_to_schedule: dict,
) -> dict:
    """讀日校某年課程節數預估表 + 公式，回傳結構化 schedule（供 Phase 2.2 編輯器用）。

    動態從 sum_up_col 掃 SUM 公式 → 每個 domain 的 range。A1 完整仿 xlsx：每科目
    列同時存上/下學期加權公式（含每班係數），讓前端可在編輯 cell 後即時重算 subject
    總節數並回推 domain total。

    summary_to_schedule: {sum_row_in_day_sheet: {'domain_id', 'summary_row'}}。同一個 domain
    可能被多個 sum_row 組成（例：health_career 是 AH125 + AH127），共用相同 domain_id。
    """
    max_col = layout["max_class_col"]
    up_col = layout["weighted_up_col"]
    down_col = layout["weighted_down_col"]
    sum_up_col = layout["sum_up_col"]
    sum_down_col = layout["sum_down_col"]
    up_letter = openpyxl.utils.get_column_letter(up_col)
    down_letter = openpyxl.utils.get_column_letter(down_col)
    sum_up_letter = openpyxl.utils.get_column_letter(sum_up_col)
    sum_down_letter = openpyxl.utils.get_column_letter(sum_down_col)

    class_columns = []
    for c in range(4, max_col + 1):
        letter = openpyxl.utils.get_column_letter(c)
        dept = ws_v.cell(row=3, column=c).value
        semester = ws_v.cell(row=4, column=c).value
        class_columns.append({
            "col": letter,
            "col_index": c,
            "dept": str(dept) if dept is not None else None,
            "semester": str(semester) if semester is not None else None,
        })

    # 掃 sum_up_col 裡所有 =SUM(...) 找 domain range；排除全 sheet grand total（跨>100 行）
    sum_blocks: list[dict] = []
    max_scan = ws_f.max_row
    for r in range(4, max_scan + 1):
        fval = ws_f.cell(row=r, column=sum_up_col).value
        if not isinstance(fval, str) or not fval.startswith("="):
            continue
        parsed = _parse_sum_range(fval)
        if not parsed or parsed["col"] != up_letter:
            continue
        if parsed["end"] - parsed["start"] > 100:
            continue
        # 對應的下學期 SUM（同列、sum_down_col）
        down_f = ws_f.cell(row=r, column=sum_down_col).value
        down_parsed = _parse_sum_range(down_f) if isinstance(down_f, str) else None
        sum_blocks.append({
            "sum_row": r,
            "up_formula": fval,
            "up_range": parsed,
            "down_formula": down_f if isinstance(down_f, str) else None,
            "down_range": down_parsed,
            "value_up":   ws_v.cell(row=r, column=sum_up_col).value,
            "value_down": ws_v.cell(row=r, column=sum_down_col).value,
        })

    def _extract_cell_formula(row: int, col_idx: int) -> dict:
        """抽加權列的公式或 raw 值。回傳 {formula, coef, const, value, raw_value}。"""
        fval = ws_f.cell(row=row, column=col_idx).value
        vval = ws_v.cell(row=row, column=col_idx).value
        int_val = int(vval) if isinstance(vval, (int, float)) else 0
        if isinstance(fval, str) and fval.startswith("="):
            parsed = _parse_weighted_formula(fval)
            return {
                "formula":   fval,
                "coef":      parsed["coef"] if parsed else {},
                "const":     parsed["const"] if parsed else 0,
                "value":     int_val,
                "raw_value": None,
            }
        raw = int(fval) if isinstance(fval, (int, float)) else int_val
        return {
            "formula":   None,
            "coef":      {},
            "const":     0,
            "value":     int_val,
            "raw_value": raw,
        }

    domains: list[dict] = []
    for blk in sum_blocks:
        rng = blk["up_range"]
        r_start = rng["start"]
        r_end = rng["end"]
        mapping = summary_to_schedule.get(blk["sum_row"], {})
        domain_id = mapping.get("domain_id")
        summary_row = mapping.get("summary_row")

        subjects = []
        first_subject_name: str | None = None
        for r in range(r_start, r_end + 1):
            name = ws_v.cell(row=r, column=3).value
            if name is None:
                continue
            if first_subject_name is None:
                first_subject_name = str(name)
            cells = {}
            for col_spec in class_columns:
                v = ws_v.cell(row=r, column=col_spec["col_index"]).value
                if isinstance(v, (int, float)):
                    cells[col_spec["col"]] = int(v)
            subjects.append({
                "row":           r,
                "name":          str(name),
                "cells":         cells,
                "weighted_up":   _extract_cell_formula(r, up_col),
                "weighted_down": _extract_cell_formula(r, down_col),
            })

        # 若 summary 沒參照到這個 sum_row（如 national_defense 公式 '=9+進113!J73' 未引用 day sheet），
        # 退而用第一個科目名對 DOMAIN_IDS 的中文名模糊比對
        if domain_id is None and first_subject_name:
            for sr, (did, zh) in DOMAIN_IDS.items():
                if first_subject_name.strip() in zh or zh in first_subject_name.strip():
                    domain_id = did
                    summary_row = sr
                    break

        domains.append({
            "domain_id":   domain_id,
            "summary_row": summary_row,
            "sum_row":     blk["sum_row"],
            "sum_up": {
                "col":     sum_up_letter,
                "formula": blk["up_formula"],
                "range":   rng,
                "value":   int(blk["value_up"]) if isinstance(blk["value_up"], (int, float)) else 0,
            },
            "sum_down": {
                "col":     sum_down_letter,
                "formula": blk["down_formula"],
                "range":   blk["down_range"],
                "value":   int(blk["value_down"]) if isinstance(blk["value_down"], (int, float)) else 0,
            },
            "subjects": subjects,
        })

    return {
        "class_columns": class_columns,
        "layout": {
            "max_class_col":    openpyxl.utils.get_column_letter(max_col),
            "weighted_up_col":  up_letter,
            "weighted_down_col":down_letter,
            "sum_up_col":       sum_up_letter,
            "sum_down_col":     sum_down_letter,
        },
        "domains": domains,
    }


def _map_summary_to_schedule(ws_summary: Any) -> dict:
    """掃 summary!D/L 公式，抽 '日113!AH5' 那種 ref，建立 {year: {sum_row: {domain_id, summary_row}}}。

    key 是 day sheet 裡 SUM 所在的那一行（例：113 的 AH5、114 的 AJ5）。同一個 domain（如
    health_career）可能在公式裡同時參照到兩個 sum_row（AH125 + AH127），則兩筆都對回同個
    domain_id。下學期 domain sum（AI/AK）的行號跟上學期是同一行，所以也可以 sum_row 視之。
    """
    result: dict[str, dict[int, dict]] = {"113": {}, "114": {}}
    # summary 行 → domain_id
    for summary_row, (domain_id, _) in DOMAIN_IDS.items():
        # D 欄 = 113 需求公式；L 欄 = 114 需求公式
        for year, col_idx in (("113", 4), ("114", 12)):
            f = ws_summary.cell(row=summary_row, column=col_idx).value
            if not isinstance(f, str):
                continue
            # 抓所有 '日<YYY>!<COL><ROW>'；中文 sheet 名含在 '日113課程...' 等
            for m in re.finditer(r"日\s*(\d{3})[^!]*!\s*([A-Z]+)(\d+)", f):
                ref_year = m.group(1)
                ref_row = int(m.group(3))
                if ref_year != year:
                    continue
                # 同 sum_row 可能對到多個 domain（理論上不會，但保底：後到覆蓋先到，或維持 list）
                # 這裡選「先到為準」，因為 summary 列順序 = domain 自然順序
                result[year].setdefault(ref_row, {
                    "domain_id":   domain_id,
                    "summary_row": summary_row,
                })
    return result


def _extract_evening_schedule(ws_v: Any) -> dict:
    """讀進113課程節數一覽表（7 班、無學期拆分）。"""
    class_columns = []
    for c in range(3, 10):  # C..I
        letter = openpyxl.utils.get_column_letter(c)
        grade = ws_v.cell(row=2, column=c).value
        dept = ws_v.cell(row=3, column=c).value
        class_columns.append({
            "col": letter,
            "col_index": c,
            "grade": str(grade) if grade is not None else None,
            "dept": str(dept) if dept is not None else None,
        })
    # 領域列：A 欄有字串，直到下一個 A 欄有字串
    sections = []
    current = None
    for r in range(4, ws_v.max_row + 1):
        a = ws_v.cell(row=r, column=1).value
        b = ws_v.cell(row=r, column=2).value
        if a:
            if current:
                sections.append(current)
            current = {
                "section_key": str(a).strip().replace("\n", ""),
                "start_row": r,
                "subjects": [],
            }
        if current is None:
            continue
        if b:
            cells = {}
            for col_spec in class_columns:
                v = ws_v.cell(row=r, column=col_spec["col_index"]).value
                if isinstance(v, (int, float)):
                    cells[col_spec["col"]] = int(v)
            current["subjects"].append({
                "row": r,
                "name": str(b).strip(),
                "cells": cells,
                "subtotal_J": ws_v.cell(row=r, column=10).value,
                "total_K":    ws_v.cell(row=r, column=11).value,
            })
    if current:
        sections.append(current)
    return {
        "class_columns": class_columns,
        "sections": sections,
    }


def _extract_evening_baseline(ws_v: Any) -> dict:
    """讀 sheet 5 進修部基本節數比較表 (B=基本, C=分配)。僅用於備查，不影響計算。"""
    out = []
    for r in range(5, ws_v.max_row + 1):
        a = ws_v.cell(row=r, column=1).value
        if not a:
            continue
        b = ws_v.cell(row=r, column=2).value
        c = ws_v.cell(row=r, column=3).value
        d = ws_v.cell(row=r, column=4).value
        if b is None and c is None:
            continue
        out.append({
            "group": str(a).strip().replace("\n", ""),
            "baseline_B": b,
            "allocated_C": c,
            "remainder_D": d,
        })
    return {"rows": out}


def parse_excel(xlsx_path: str) -> dict:
    wb_v = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_f = openpyxl.load_workbook(xlsx_path, data_only=False)

    ws_v = wb_v[wb_v.sheetnames[SUMMARY_SHEET_INDEX]]
    ws_f = wb_f[wb_f.sheetnames[SUMMARY_SHEET_INDEX]]

    day_113_name = wb_v.sheetnames[DAY_113_SHEET_INDEX]
    day_114_name = wb_v.sheetnames[DAY_114_SHEET_INDEX]
    evening_113_name = wb_v.sheetnames[EVENING_113_SHEET_INDEX]
    evening_baseline_name = wb_v.sheetnames[EVENING_BASELINE_SHEET_INDEX]

    required_breakdown: dict[str, dict[str, dict]] = {"113": {}, "114": {}}

    domains: list[dict] = []

    for row, (dom_id, dom_name) in DOMAIN_IDS.items():
        d_formula = _extract_formula(ws_f.cell(row=row, column=4).value)   # D 欄 113 需求
        l_formula = _extract_formula(ws_f.cell(row=row, column=12).value)  # L 欄 114 需求
        required_breakdown_113 = _decompose_required_formula(
            "=" + d_formula if d_formula else None,
            wb_v, day_113_name, evening_113_name,
        )
        required_breakdown_114 = _decompose_required_formula(
            "=" + l_formula if l_formula else None,
            wb_v, day_114_name, evening_113_name,
        )
        # 若為常數格（例：全民國防 D=11 無公式），退回硬編碼值
        if not d_formula:
            raw_d = ws_v.cell(row=row, column=4).value
            if isinstance(raw_d, (int, float)):
                required_breakdown_113 = {"day": int(raw_d), "evening": 0, "total": int(raw_d), "source": None}
        if not l_formula:
            raw_l = ws_v.cell(row=row, column=12).value
            if isinstance(raw_l, (int, float)):
                required_breakdown_114 = {"day": int(raw_l), "evening": 0, "total": int(raw_l), "source": None}
        required_breakdown["113"][dom_id] = required_breakdown_113
        required_breakdown["114"][dom_id] = required_breakdown_114
        positions: dict[str, int] = {}
        for spec in POSITION_COLUMNS:
            col_idx = openpyxl.utils.column_index_from_string(spec["col"])
            positions[spec["key"]] = _safe_int(ws_v.cell(row=row, column=col_idx).value)

        substitute_count = _safe_int(ws_v.cell(row=row, column=37).value)   # AK
        substitute_evening_count = _safe_int(ws_v.cell(row=row, column=39).value)  # AM

        c_formula = _extract_formula(ws_f.cell(row=row, column=3).value)
        g_formula = _extract_formula(ws_f.cell(row=row, column=7).value)
        o_formula = _extract_formula(ws_f.cell(row=row, column=15).value)   # O 欄：115 未來預估

        ah_formula = _extract_formula(ws_f.cell(row=row, column=34).value)
        ai_formula = _extract_formula(ws_f.cell(row=row, column=35).value)
        ah_cell    = ws_v.cell(row=row, column=34).value
        ai_cell    = ws_v.cell(row=row, column=35).value

        ah_override   = _extract_ah_override(ah_cell, ah_formula)
        ai_override   = _extract_ai_override(ai_cell, ai_formula)
        ai_adjustment = _parse_ai_adjustment(ai_formula)

        c_constant = None
        if c_formula is None:
            c_constant = ws_v.cell(row=row, column=3).value
        g_constant = None
        if g_formula is None:
            g_constant = ws_v.cell(row=row, column=7).value
        o_constant = None
        if o_formula is None:
            o_constant = ws_v.cell(row=row, column=15).value

        substitute_note = ws_v.cell(row=row, column=36).value   # AJ
        name_list       = ws_v.cell(row=row, column=38).value   # AL
        remark_person   = ws_v.cell(row=row, column=40).value   # AN
        remark_event    = ws_v.cell(row=row, column=41).value   # AO
        future_note     = ws_v.cell(row=row, column=45).value   # AS

        excel_raw = {
            "113": {
                "C":  ws_v.cell(row=row, column=3).value,
                "D":  ws_v.cell(row=row, column=4).value,
                "E":  ws_v.cell(row=row, column=5).value,
                "G":  ws_v.cell(row=row, column=7).value,
                "H":  ws_v.cell(row=row, column=8).value,
                "I":  ws_v.cell(row=row, column=9).value,
                "AH": ws_v.cell(row=row, column=34).value,
                "AI": ws_v.cell(row=row, column=35).value,
                "AK": ws_v.cell(row=row, column=37).value,
            },
            "114": {
                "K": ws_v.cell(row=row, column=11).value,
                "L": ws_v.cell(row=row, column=12).value,
                "M": ws_v.cell(row=row, column=13).value,
            },
            "115_future": {
                "O": ws_v.cell(row=row, column=15).value,
                "P": ws_v.cell(row=row, column=16).value,
                "Q": ws_v.cell(row=row, column=17).value,
            },
        }

        required_113 = ws_v.cell(row=row, column=4).value
        required_114 = ws_v.cell(row=row, column=12).value

        domains.append({
            "id": dom_id,
            "name": dom_name,
            "summary_row": row,
            "positions": positions,
            "substitute_count": substitute_count,
            "substitute_evening_count": substitute_evening_count,
            "base_homeroom_formula": c_formula,
            "base_homeroom_constant": c_constant,
            "base_position_formula": g_formula,
            "base_position_constant": g_constant,
            "future_base_formula": o_formula,
            "future_base_constant": o_constant,
            "ah_override":   ah_override,
            "ai_override":   ai_override,
            "ai_adjustment": ai_adjustment,
            "required_hours": {
                "113": _safe_int(required_113),
                "114": _safe_int(required_114),
            },
            "substitute_note":  substitute_note or "",
            "name_list":        name_list       or "",
            "remark_person":    remark_person   or "",
            "remark_event":     remark_event    or "",
            "future_note":      future_note     or "",
            "_excel_raw":       excel_raw,
        })

    # 114 學年教師編制目前未在 Excel 中獨立記錄，先沿用 113；教務主任之後可獨立編輯
    def _year_state(dom, year_key):
        br = required_breakdown[year_key][dom["id"]]
        return {
            "positions": dict(dom["positions"]),
            "substitute_count": dom["substitute_count"],
            "substitute_evening_count": dom["substitute_evening_count"],
            "required_hours": dom["required_hours"][year_key],
            "required_day":      br["day"],
            "required_evening":  br["evening"],
            "required_total":    br["total"],
            "required_formula":  br["source"],
        }
    years_data = {
        "113": {dom["id"]: _year_state(dom, "113") for dom in domains},
        "114": {dom["id"]: _year_state(dom, "114") for dom in domains},
    }

    # 驗證：required_total 必須 == 原 required_hours（= sheet0 D/L 欄值）
    for y in ("113", "114"):
        for dom in domains:
            br = required_breakdown[y][dom["id"]]
            expected = dom["required_hours"][y]
            if br["total"] != expected:
                raise ValueError(
                    f"required_total 與 Excel {y} {dom['id']} 不符: "
                    f"parsed={br['total']} excel={expected} formula={br['source']!r}"
                )

    # summary 公式決定 day sheet 哪個 sum_row 對應哪個 domain
    summary_to_schedule = _map_summary_to_schedule(ws_f)

    # 課程節數表（schedule）— Phase 2.2 編輯器將讀寫此結構
    schedule = {
        "113": {
            "day":     _extract_day_schedule(
                wb_v[day_113_name], wb_f[day_113_name],
                DAY_SHEET_LAYOUTS["113"], summary_to_schedule["113"],
            ),
            "evening": _extract_evening_schedule(wb_v[evening_113_name]),
            "evening_baseline": _extract_evening_baseline(wb_v[evening_baseline_name]),
        },
        "114": {
            "day":     _extract_day_schedule(
                wb_v[day_114_name], wb_f[day_114_name],
                DAY_SHEET_LAYOUTS["114"], summary_to_schedule["114"],
            ),
            "evening": None,   # Excel 目前僅有 進113 一份；114 進修部沿用 113
            "evening_baseline": None,
        },
    }

    domain_meta = []
    for dom in domains:
        domain_meta.append({
            "id": dom["id"],
            "name": dom["name"],
            "summary_row": dom["summary_row"],
            "base_homeroom_formula": dom["base_homeroom_formula"],
            "base_homeroom_constant": dom["base_homeroom_constant"],
            "base_position_formula": dom["base_position_formula"],
            "base_position_constant": dom["base_position_constant"],
            "future_base_formula":  dom["future_base_formula"],
            "future_base_constant": dom["future_base_constant"],
            "ah_override":   dom["ah_override"],
            "ai_override":   dom["ai_override"],
            "ai_adjustment": dom["ai_adjustment"],
            "substitute_note": dom["substitute_note"],
            "name_list": dom["name_list"],
            "remark_person": dom["remark_person"],
            "remark_event": dom["remark_event"],
            "future_note": dom["future_note"],
        })

    return {
        "schema_version": 2,
        "source": "teacher.xlsx",
        "default_year": "113",
        "available_years": ["113", "114"],
        "position_columns": POSITION_COLUMNS,
        "domains": domain_meta,
        "years": years_data,
        "schedule": schedule,
        "excel_raw_snapshot": {dom["id"]: dom["_excel_raw"] for dom in domains},
    }


def main() -> None:
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    xlsx = os.path.join(here, "data", "teacher.xlsx")
    out  = os.path.join(here, "data", "school_data.json")

    data = parse_excel(xlsx)

    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"OK  寫入 {out}")
    print(f"    領域數 = {len(data['domains'])}, 年度 = {data['available_years']}")


if __name__ == "__main__":
    main()
